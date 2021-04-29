import json
import math
import openpyxl
from pathlib import Path


# WorkspaceProperties
directoryId = ""
bundleId = ""

volumeEncryptionKey = ""
userVolumeEncryptionEnabled = False
rootVolumeEncryptionEnabled = False
runningMode = "ALWAYS_ON"
runningModeAutoStopTimeoutInMinutes = 60
rootVolumeSizeGib = 80
userVolumeSizeGib = 10
computeTypeName = "PERFORMANCE"


# Tags
user_key = ""
user_value = ""


# AWS Configs
runningModes = ["ALWAYS_ON", "AUTO_STOP"]
computeTypeNames = ["VALUE", "STANDARD", "PERFORMANCE", "POWER", "GRAPHICS", "POWERPRO", "GRAPHICSPRO"]


# Process config
names = []
data = []


# CLIENT DATABASE -> CLIENT: [[DIRECTORIES], [BUNDLES]]
clients = {
    'client1': [["dir1", "dir2"], ["img1", "img2", "img3", "img4", "img5"]],
    'client2': [["dir1", "dir2"], ["img1", "img2", "img3", "img4"]],
    'client3': [["dir1", "dir2"], ["img1"]],
    'client4': [["dir1", "dir2"], ["img1", "img2", "img3"]],
    'client5': [["dir1", "dir2"], ["img1", "img2"]],
    'client6': [["dir1", "dir2"], ["img1", "img2", "img3"]],
    'client7': [["dir1", "dir2"], ["img1", "img2"]],
    'client8': [["dir1", "dir2"], ["img1", "img2", "img3"]],
    'client9': [["dir1", "dir2"], ["img1", "img2", "img3", "img4"]]
}


# Print available resources, select needed
def print_and_assign(available):
    for i in range(len(available)):
        print(str(i + 1) + ". " + available[i])
    selected = input("Select option: ")
    selected = available[int(selected) - 1]
    print()
    return selected


# Open xl input
xlsx_file = Path('', 'user_list.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file)
sheet = wb_obj.active
n_jsons = math.ceil(sheet.max_row / 25)

# Append users into names[]:
for row in sheet.iter_rows(sheet.max_column):
    for cell in row:
        cell_value = cell.value
        cell_value.replace(" ", "")
        names.append(cell_value)
user_count = len(names)


# Create JSONs to upload via CLI function
def process():
    aws_max_workload = 25
    for g in range(n_jsons):
        if aws_max_workload > len(names):
            aws_max_workload = len(names)
        for h in range(aws_max_workload):
            data.append({
                'DirectoryId': directoryId,
                'UserName': names[h],
                'BundleId': bundleId,
                'VolumeEncryptionKey': volumeEncryptionKey,
                'UserVolumeEncryptionEnabled': userVolumeEncryptionEnabled,
                'RootVolumeEncryptionEnabled': rootVolumeEncryptionEnabled,
                'WorkspaceProperties': {
                    'RunningMode': runningMode,
                    'RunningModeAutoStopTimeoutInMinutes': runningModeAutoStopTimeoutInMinutes,
                    'RootVolumeSizeGib': rootVolumeSizeGib,
                    'UserVolumeSizeGib': userVolumeSizeGib,
                    'ComputeTypeName': computeTypeName
                },
                'Tags': [
                    {
                        'Key': user_key,
                        'Value': user_value
                    },
                ]
            })

        with open('workload' + str(g) + '.json', 'w') as outfile:
            json.dump(data, outfile)

        for index in range(aws_max_workload):
            names.pop(0)

        data.clear()

    return True

# #########################################################################


# UI Start
print()

print("***aws_wsAutoDeployV1.0***")
print(".-._.-._.W311C0M3._.-._.-.")
print("***aws_wsAutoDeployV1.0***")

print()
print()

print("You are about to create " + str(user_count) + " workspaces.")
print()

converted_list = []
for key in clients.keys():
    converted_list.append(key)
print("Clients:  " + str(converted_list))
print()

activeClient = input("Type client name:")
currentClientConfig = clients.get(activeClient)
print("Active client is: " + activeClient)

# Assign lists to individual variables to manage later
working_directories = currentClientConfig[0]
available_bundles = currentClientConfig[1]


print()
print("Available directories: ")
directoryId = print_and_assign(working_directories)
print("Working directory: " + directoryId)

print()
print("Available OS bundles: ")
bundleId = print_and_assign(available_bundles)
print("Selected bundle: " + bundleId)

print()
print("Volume encryption is set to FALSE." + str(userVolumeEncryptionEnabled))
if input("Is this parameter correct?(y/n): ") == "n":
    userVolumeEncryptionEnabled = True
    rootVolumeEncryptionEnabled = True
    volumeEncryptionKey = input("Insert encryption key: ")

print()
print("The running mode is set to: " + runningMode)
if input("Is this parameter correct?(y/n): ") == "n":
    runningMode = print_and_assign(runningModes)
    print("Running mode: " + runningMode)

print()
print("The compute type is set to: " + computeTypeName)
if input("Is this parameter correct?(y/n): ") == "n":
    computeTypeName = print_and_assign(computeTypeNames)
    print("Compute type: " + computeTypeName)

print()
print("The instances default storage is set to: " + str(rootVolumeSizeGib) + "GB root, " + str(userVolumeSizeGib) + "GB user")
if input("Is this parameter correct?(y/n): ") == "n":
    rootVolumeSizeGib = input("Assign amount of Gb root: ")
    userVolumeSizeGib = input("Assign amount of Gb user: ")


print()
print("     This are the current configurations:")
print("     Directory: " + directoryId)
print("     BundleId: " + bundleId)
print()

print("     Volume key: " + volumeEncryptionKey)
print("     User encryption: " + str(userVolumeEncryptionEnabled))
print("     Root encryption: " + str(rootVolumeEncryptionEnabled))
print()

print("     Running mode: " + runningMode)
print("     Idle stop time: " + str(runningModeAutoStopTimeoutInMinutes))
print("     Root Volume: " + str(rootVolumeSizeGib) + "Gb")
print("     User Volume: " + str(userVolumeSizeGib) + "Gb")
print("     Compute type: " + computeTypeName)
print()


if input("Do you wish to proceed creating " + str(n_jsons) + "JSON files ?(y/n) ") == "y":
    print("Creating files...")

    process()

    print("*********************************************")
    print("************ OPERATION SUCCESS **************")
    print("*********************************************")
    print("************* OUTPUT " + str(n_jsons) + " .JSONs ***************")
    print("*********************************************")
    print("************ OPERACION EXITOSA **************")
    print("*********************************************")

else:
    print("Operation canceled. Execute again.")
    exit()

